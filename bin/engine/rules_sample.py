
from util.logger import error
import values as v
import logging
from openpyxl import load_workbook

def get_sample_type(pt_spl_type, state):
    logging.info('Loading sample type map...')
    # TODO optimize this later
    # I don't like that I load the map workbook every call. It's small,
    # but it'd still be better to only do this once.
    wb_map_spl_type = None # do I need this?
    try: wb_map_spl_type = load_workbook(v.path_map_sample_types, read_only=True) # must close manually if read_only=True
    except FileNotFoundError:
        error('MISSING SAMPLE TYPE MAPPING!!! ABORTING!!! Expected in: ' + v.path_map_sample_types)
        return None

    try:
        map = wb_map_spl_type[state]
        logging.info('Overriding STI sample type map with state rules: ' + state)
    except:
        try:
            map = wb_map_spl_type[v.default]
            logging.info('Default STI sample type map loaded.')
        except:
            wb_map_spl_type.close()
            error('COULDN\'T LOAD DEFAULT STI SAMPLE TYPE MAP!!! ' + v.path_map_sample_types)
            return None

    st_tuple = None
    # +1 skip header, +1 include last row
    for row in range(map.min_row + 1, map.max_row + 1):
        r = str(row)
        spl_type = map[v.map_st_col_code + r].value
        if not spl_type or spl_type.lower() != pt_spl_type.lower(): continue # TODO add logging?
        logging.info('Matched patient ' + pt_spl_type + ' to map ' + spl_type)

        name = map[v.map_st_col_name + r].value
        display = map[v.map_st_col_display + r].value
        form = map[v.map_st_col_form + r].value
        st_tuple = (spl_type, name, display, form)

        for item in st_tuple:
            if not item or item == '':
                error('SAMPLE TYPE MAPPING FILE IMPROPERLY CONFIGURED!!! ' + v.path_map_sample_types)
                return None

    wb_map_spl_type.close()
    if not st_tuple: error('COULD NOT MATCH SAMPLE TYPE FROM MAPPING FILE!!! ' + v.path_map_sample_types)
    return st_tuple
