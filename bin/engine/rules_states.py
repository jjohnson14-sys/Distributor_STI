
from engine.rules_sti import get_generic_code
from util.logger import error, warning
import values as v

import logging
from os import mkdir
from os.path import exists, join
from openpyxl import load_workbook

from alive_progress import alive_it

i_county, i_markttl = 0, 1
i_target = 0
rule_tuple_reports_per = 1

# rules struct: { state: { ( county, STI ): target } }
class Rules_States:
    def __init__(self, state_set):
        wb_rules = None
        if not exists(v.dir_rules):
            error('NO RULES DIRECTORY!!! PLEASE ADD ' + v.file_rules_state + ' TO ' + v.dir_rules + ' IN THE PROGRAM DIRECTORY!!!')
            try: mkdir(join(v.cwd, v.dir_rules))
            except PermissionError: error('NO PERMISSIONS TO MAKE ' + v.dir_rules + ' AUTOMATICALLY!!!')
            return
        try: wb_rules = load_workbook(filename=v.path_rules_state, read_only=True)
        except FileNotFoundError:
            error('MISSING STATE RULES FILE!!! EXPECTED LOCATION: ' + v.path_rules_state)
            return

        print('Loading state rules...')
        logging.info('Loading rules for states: ' + str(state_set))
        rules = {}
        for state in alive_it(state_set): # load data from relevant states
            if state in wb_rules.sheetnames:
                s_rules = self.load_state(wb_rules[state])
                if s_rules == False: continue # error captured in load_state()
                rules = rules | {state: s_rules}
            else: warning('No rules detected for state ' + state + '! Skipping all positives from ' + state + '.')
        self.city_dict = self.load_cities(wb_rules)
        wb_rules.close()
        del wb_rules # probably unnecessary now that DL happens before this lul

        self.rules_dict = rules # TODO find out if it's better to dict() or not to dict()
        logging.info('Loaded state rules.')

    # TODO TODO TODO rewrite this and get_doh_target into one method if you have time
    def get_city_doh_target(self, state, county, sti, pt_city):
        city_dict = self.city_dict.get(state)
        options = { # { (city, sti): (target, template, reports_per) }
            r_city_sti[1]: city_dict.get(r_city_sti)
            for r_city_sti in city_dict
            if r_city_sti[0].lower() == pt_city.lower()
        }

        tally = 0 # confirm only one ! is used per county
        for option in options:
            if v.flag_not in option:
                tally += 1
                if tally > 1:
                    error('IMPROPER USAGE OF FLAG "!" DETECTED!!! THERE CAN BE ONLY ONE!!!')
                    logging.info('Skipping...')
                    return None
                if option.index(v.flag_not) != 0:
                    error('MISCONFIGURED "!" FLAG!!!')
                    logging.info('Skipping...')
                    return None

        sti = get_generic_code(sti)
        # first: find a direct match if possible
        # option[0] = county | [1] = sti | same as it ever was
        for option in options:
            if sti.lower() != option.lower(): continue
            target = options.get(sti)#[i_target]
            logging.info('Direct rule match found:')
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No direct rule match found.')

        # second, evaluate not (!)
        for option in options:
            #if v.rules_flag_not not in option or sti.lower() == option[1:].lower(): continue
            if v.flag_not not in option: continue
            if sti.lower() == option[1:].lower(): continue
            target = options.get(option)#[i_target]
            logging.info('NOT rule match found:')
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No NOT (' + v.flag_not + ') rule match found.')

        for option in options:
            if option.lower() != v.flag_all.lower(): continue
            target = options.get(option)#[i_target]
            logging.info('ALL rule match found:')
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No ALL rule match found.')

        return None

    # TODO this whole block should be rewritten and optimized
    # there shouldn't be any error handling necessary for rules in this function
    # TODO confirm / list return types
    # return false      report must be sent manually
    # return none       something went wrong upstream
    def get_doh_target(self, state, county, sti, pt_city):
        if county is None: county = 'None'
        blurb = 'state: ' + state + ', county: ' + county
        s_blurb = blurb + ', sti: ' + sti
        logging.info('Locating DOH target for ' + s_blurb)
        state_dict = self.rules_dict.get(state)
        if not state_dict:
            warning('Somehow, no rules for ' + state + ' were loaded. Manual intervention required.')
            return None # should never make it here, in theory

        # check cities
        if state in self.city_dict:
            city_result = self.get_city_doh_target(state, county, sti, pt_city)
            if city_result is not None: return city_result

        options = {
            r_county_sti[1]: state_dict.get(r_county_sti)
            for r_county_sti in state_dict
            if r_county_sti[0].lower() == county.lower()
            or r_county_sti[0].lower() == v.flag_all.lower()
        }

        tally = 0 # confirm only one ! is used per county
        for option in options:
            if v.flag_not in option:
                tally += 1
                if tally > 1:
                    error('IMPROPER USAGE OF FLAG "!" DETECTED!!! THERE CAN BE ONLY ONE: ' + blurb + ', rules: ' + str(options))
                    logging.info('Skipping...')
                    return None
                if option.index(v.flag_not) != 0:
                    error('MISCONFIGURED "!" FLAG!!! ' + blurb + ', rule: ' + option)
                    logging.info('Skipping...')
                    return None

        sti = get_generic_code(sti)
        # first: find a direct match if possible
        # option[0] = county | [1] = sti | same as it ever was
        for option in options:
            if sti.lower() != option.lower(): continue
            target = options.get(sti)#[i_target]
            logging.info('Direct rule match found:')
            logging.info('Csv info: ' + s_blurb)
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No direct rule match found.')

        # second, evaluate not (!)
        for option in options:
            #if v.rules_flag_not not in option or sti.lower() == option[1:].lower(): continue
            if v.flag_not not in option: continue
            if sti.lower() == option[1:].lower(): continue
            target = options.get(option)#[i_target]
            logging.info('NOT rule match found:')
            logging.info('Csv info: ' + s_blurb)
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No NOT (' + v.flag_not + ') rule match found.')

        for option in options:
            if option.lower() != v.flag_all.lower(): continue
            target = options.get(option)#[i_target]
            logging.info('ALL rule match found:')
            logging.info('Csv info: ' + s_blurb)
            logging.info('Rule: ' + option + ', target: ' + str(target))
            return target
        logging.info('No ALL rule match found.')

        logging.info('Manual sending required for ' + s_blurb)
        return False

    def load_cities(self, wb_rules):
        logging.info('Loading city rules...')
        sheet_city = None
        try: sheet_city = wb_rules[v.sheet_city]
        except:
            error('COULD NOT READ DATA FOR CITY-BASED RULES!!!')
            return None

        city_dict = {}
        for row in range(sheet_city.min_row + 1, sheet_city.max_row + 1):
            str_r = str(row)
            cell_st = v.rules_col_county + str_r # col county == col city
            cell_sti = v.rules_col_sti + str_r
            cell_target = v.rules_col_doh_target + str_r
            cell_template = v.rules_col_template + str_r
            cell_city = v.rules_col_city + str_r

            state = sheet_city[cell_st].value
            city = sheet_city[cell_city].value
            if state is None or state == '' or city is None or city == '': continue

            rule_sti = sheet_city[cell_sti].value
            if rule_sti is None or rule_sti.strip() == '':
                warning('City rule missing STI reference, state: ' + state + ', cell: ' + cell_sti + ', city: ' + city)
                continue

            l_sti = rule_sti.split(',')
            stis = [r.strip(', ') for r in l_sti] if ',' in rule_sti else [rule_sti]

            target = sheet_city[cell_target].value
            if target is None or target.strip() == '':
                target = v.flag_manual
                warning('Overriding blank target with MANUAL flag, state: ' + state + ', cell: ' + cell_sti + ', city: ' + city)
            
            template = sheet_city[cell_template].value
            if template is None:
                template = state + v.ext_sti_template
            template = join(v.dir_report_templates, template)

            reports_per = sheet_city[v.rules_col_reports_per + str_r].value
            for sti in stis:
                city_sti = (city, sti)
                if city_sti in city_dict:
                    error('DUPLICATE CITY STI RULE DETECTED!!! THIS CONFLICT MUST BE FIXED BEFORE STI REPORTS FOR THIS CITY CAN BE GENERATED!!! State: ' + state + ', city: ' + city + ', STI: ' + rule_sti)
                    return False
                reports_per = reports_per.lower()
                target_tuple = (target, template, reports_per)
                city_dict.update({state: {city_sti: target_tuple}})
            
        if len(city_dict) < 1:
            warning('No valid rules detected for cities.')
            return False
        return city_dict

    def load_state(self, sheet_st):
        state = sheet_st.title
        logging.info('Loading ' + state + ' rules...')
        state_dict = {}
        # +1 to skip header, +1 to include last filled row in eval
        for row in range(sheet_st.min_row + 1, sheet_st.max_row + 1):
            str_r = str(row)
            cell_county = v.rules_col_county + str_r
            cell_sti = v.rules_col_sti + str_r
            cell_target = v.rules_col_doh_target + str_r
            cell_template = v.rules_col_template + str_r

            county = sheet_st[cell_county].value
            if county is None: continue # skip blanks

            # can't process a target if there is no STI mapped to it
            rule_sti = sheet_st[cell_sti].value
            if rule_sti is None or rule_sti.strip() == '':
                warning('County missing an STI reference, state: ' + state + ', cell: ' + cell_sti + ', county: ' + county)
                continue

            l_sti = rule_sti.split(',') # FIXME FIXME FIXME this doesn't handle '!'
            stis = [r.strip(', ') for r in l_sti] if ',' in rule_sti else [rule_sti]

            target = sheet_st[cell_target].value
            if target is None or target.strip() == '':
                target = v.flag_manual
                warning('Overriding blank target with MANUAL flag, state: ' + state + ' cell: ' + cell_target + ', county: ' + county)

            template = sheet_st[cell_template].value
            if template is None:
                template = state + v.ext_sti_template
            template = join(v.dir_report_templates, template)

            reports_per = sheet_st[v.rules_col_reports_per + str_r].value
            for sti in stis:
                county_sti = (county, sti)
                if county_sti in state_dict:
                    error('DUPLICATE STI RULE DETECTED!!! THIS CONFLICT MUST BE FIXED BEFORE STI REPORTS FOR THIS STATE CAN BE GENERATED!!! State: ' + state + ', county: ' + county + ', STI: ' + rule_sti)
                    return False
                reports_per = reports_per.lower()
                target_tuple = (target, template, reports_per)
                state_dict.update({county_sti: target_tuple})
        
        if len(state_dict) < 1:
            error('NO VALID RULES DETECTED FOR STATE ' + state + '!!!')
            return False
        return state_dict

    def has_rules(self, doh_state): return doh_state in self.rules_dict

    def get_rules(self):
        try: return self.rules_dict
        except AttributeError: return None # error message captured in constructor
    
    def get_state_rules(self, state):
        return self.rules_dict.get(state)

    def needs_county(self, state):
        sr = self.get_state_rules(state)
        for rule in sr:
            if rule[i_county].lower() == v.flag_all.lower(): return False
        return True

    def state_has_sti(self, state, sti):
        sr = self.get_state_rules(state)
        if sr is None:
            logging.info('Rules for ' + state + ' not available or doesn\'t exist.')
            return
        if v.code_hiv in sti: sti = v.code_hiv
        elif v.code_syph in sti: sti = v.code_syph
        sti = sti.lower()
        for rule in sr:
            r_sti = rule[i_markttl].lower()
            if r_sti in sti: return True
            if v.flag_all.lower() == r_sti: return True
            if v.flag_not == r_sti[:1] and sti not in r_sti: return True
        return False
