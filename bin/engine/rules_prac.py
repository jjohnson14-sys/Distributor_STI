
from data.formatter import format_phone_display as ffd
from util.logger import warning
import values as v
import logging
from openpyxl import load_workbook

# prac_name comes in as "last, first", this swaps it around and removes the comma
def format_prac_name(prac_name):
    try:
        if prac_name != '' and prac_name is not None:
            return prac_name[prac_name.index(',') + 2:] + ' ' + prac_name[:prac_name.index(',')]
    except:
        warning('Prac name: ' + prac_name + ' could not be processed.')
        return None

def load_prac_rules(prc_clinic):
    try: return load_workbook(filename=v.path_rules_prac)[prc_clinic]
    except KeyError:
        logging.info('No rules to apply to clinic ' + prc_clinic + '.')
        return None

def apply_prac_rules(info):
    prac_dict = {}
    prac_name = format_prac_name(info.get(v.prac_name))
    prc_fn = info.get(v.prc_fn)
    prc_ln = info.get(v.prc_ln)

    if ((prac_name is None or prac_name == '') and
        not ((prc_fn is None or prc_fn == '') and
        (prc_ln is None or prc_ln == ''))):
        prac_name = prc_fn + ' ' + prc_ln
    elif (prc_fn is None or prc_fn == '') and (prc_ln is None or prc_fn == ''):
        prc_fn = prac_name[:prac_name.index(' ')]
        prc_ln = prac_name[prac_name.index(' ') + 1:]

    loc_addr1 = info.get(v.loc_addr1)
    loc_addr2 = info.get(v.loc_addr2)
    loc_addr_full = loc_addr1 + ' ' + loc_addr2

    prac_dict.update({
        v.prac_name: prac_name,
        v.prc_fn: prc_fn,
        v.prc_ln: prc_ln,
        v.prc_clinic: info.get(v.prc_clinic),
        v.loc_collected_name: info.get(v.prc_clinic),
        v.prc_addr: info.get(v.prc_addr),
        v.loc_addr1: loc_addr1,
        v.loc_addr2: loc_addr2,
        v.loc_addr_full: loc_addr_full,
        #v.loc_collected_addr: info.get(v.loc_addr1) + ' ' + info.get(v.loc_addr2),
        v.loc_city: info.get(v.loc_city),
        v.loc_collected_city: info.get(v.loc_city),
        v.loc_st: info.get(v.loc_st), # TODO update to loc_collected?
        v.loc_zip: info.get(v.loc_zip), #    same here?
        v.loc_phone: ffd(info.get(v.loc_phone)),
        v.loc_collected_phone: ffd(info.get(v.loc_phone)),
        v.prc_clinic_addr: info.get(v.prc_clinic) + ', ' + loc_addr_full + info.get(v.loc_city) + ', ' + info.get(v.loc_st)
    })

    clinic = prac_dict.get(v.prc_clinic)
    pr = load_prac_rules(clinic)
    if pr is None: return prac_dict
    # if there are rules for the clinic:
    for row in range(pr.min_row + 1, pr.max_row + 1):
        r = str(row)
        cell_field = pr[v.rules_prac_col_field + r]
        field_val = cell_field.value
        cell_ovrd = pr[v.rules_prac_col_ovrd + r]
        ovrd_val = cell_ovrd.value

        if field_val is None and ovrd_val is None: continue
        elif field_val is None or ovrd_val is None:
            warning('Unmapped / blank fields detected in ' + v.file_rules_prac + ': row ' + r)
            continue

        if ovrd_val.lower() == v.flag_blank.lower(): ovrd_val = ''
        prac_dict.update({field_val: ovrd_val})
        if ovrd_val == '': ovrd_val = v.flag_blank
        logging.info('Overriding ' + prac_dict.get(v.prc_clinic) + ' form field value: ' + field_val)
        logging.info('with replacement: ' + ovrd_val)

    return prac_dict
