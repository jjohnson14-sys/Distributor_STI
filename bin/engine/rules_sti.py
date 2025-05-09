
from util.logger import error
import values as v
from openpyxl import load_workbook

def convert_code_to_name(code, device):
    wb_sti = None
    try: wb_sti = load_workbook(v.path_map_sti, read_only=True)
    except Exception as e:
        for arg in e.args: error(arg)
        error('MISSING STI MAPPING FILE!!! ABORTING!!! Expected in: ' + v.path_map_sti)
        return None
    
    sheet = wb_sti[v.map_sti_map]
    for row in range(sheet.min_row + 1, sheet.max_row + 1):
        r = str(row)
        sti_device = sheet[v.map_sti_col_device + r].value
        sti_code = sheet[v.map_sti_col_code + r].value
        sti_name = sheet[v.map_sti_col_name + r].value
        if sti_code is None or sti_code == '' or sti_name is None or sti_name == '' or sti_device is None or sti_device == '':
            error(v.file_map_sti + ' MISSING MAPPING!!! Row: ' + r)
            return None
        if sti_code.lower() != code.lower() or sti_device.lower() != device.lower(): continue
        return sti_name

    error('STI NAME MAPPING FOR ' + code + ' + ' + device + ' COULD NOT BE DETERMINED!!!')
    return None

def get_generic_code(code):
    wb_sti = None
    try: wb_sti = load_workbook(v.path_map_sti, read_only=True)
    except Exception as e:
        for arg in e.args: error(arg)
        error('MISSING STI MAPPING!!! ABORTING!!! Expected in: ' + v.path_map_sti)
        return None

    sheet = wb_sti[v.map_sti_generic]
    for row in range(sheet.min_row + 1, sheet.max_row + 1):
        r = str(row)
        sti_code = sheet[v.map_sti_g_code + r].value
        if sti_code is None or sti_code == '':
            error(v.file_map_sti + ' MISSING MAPPING!!! Row: ' + r)
            return None
        if sti_code.lower() not in code.lower(): continue
        return sti_code

    error('NO GENERIC CODE MATCH FOUND FOR CODE ' + code + '!!!')
    return None


def get_generic_name(code):
    wb_sti = None
    try: wb_sti = load_workbook(v.path_map_sti, read_only=True)
    except Exception as e:
        for arg in e.args: error(arg)
        error('MISSING STI MAPPING!!! ABORTING!!! Expected in: ' + v.path_map_sti)
        return None

    sheet = wb_sti[v.map_sti_generic]
    for row in range(sheet.min_row + 1, sheet.max_row + 1):
        r = str(row)
        sti_code = sheet[v.map_sti_g_code + r].value
        sti_name = sheet[v.map_sti_g_name + r].value
        if sti_code is None or sti_code == '' or sti_name is None or sti_name == '':
            error(v.file_map_sti + ' MISSING MAPPING!!! Row: ' + r)
            return None
        if sti_code.lower() not in code.lower(): continue
        return sti_name

    error('NO GENERIC NAME MATCH FOUND FOR CODE ' + code + '!!!')
    return None